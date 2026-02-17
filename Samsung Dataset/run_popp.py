"""
Popp 256-Cell LGN Pipeline
============================
Extracts HPPC relaxation from each cell CSV, runs LGN to get τ₁, τ₂,
then tests whether τ separates SOH groups across ~400 cells.

This is the "n=256" validation that upgrades the paper.

Usage:
  python run_popp.py --data_dir popp_data/ --batch_list Batch_List.xlsx --gpu 0

Structure expected:
  popp_data/
    Engineered_Cleaned_Report_Samsung_INR21700-50E_Zelltester_MCU1_*.csv
    Engineered_Cleaned_Report_Samsung_INR21700-50E_Zelltester_MCU2_*.csv
    ...
  (across multiple batch subfolders, or all in one folder)
"""

import argparse, glob, json, os, sys, time, warnings
import numpy as np
import pandas as pd
from scipy import stats
from scipy.optimize import curve_fit

warnings.filterwarnings('ignore')


# ============================================================================
# Step 1: Extract HPPC relaxation from a single CSV
# ============================================================================

def extract_relaxation(csv_path, min_pulse_current=1.0, relax_window=None):
    """
    Extract post-discharge relaxation from Popp HPPC CSV.
    
    Protocol structure (per microcycle):
      - Rest (~10s)
      - Charge pulse +3.7A (~10s)
      - Relaxation (~35s)
      - Discharge pulse -4.9A (~10s)
      - Relaxation (~40s)  <-- TARGET
    
    Args:
        csv_path: Path to CSV file
        min_pulse_current: Minimum |I| to detect pulse (A)
        relax_window: (start_s, end_s) seconds of relaxation to use, or None for auto
    
    Returns:
        dict with 't', 'V', 'eta', 'V_end', 'cell_id', 'batch', 'mcu', 
        'discharge_I', 'soc_voltage' or None if extraction fails
    """
    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        print(f"  ERROR reading {csv_path}: {e}", flush=True)
        return None
    
    if 'I' not in df.columns or 'U' not in df.columns:
        print(f"  ERROR: Missing I or U columns in {csv_path}", flush=True)
        return None
    
    I = df['I'].values
    U = df['U'].values
    n = len(I)
    
    # Compute elapsed time (1 kHz sampling = 1ms per sample)
    t = np.arange(n) * 0.001  # seconds
    
    # Find discharge pulse: large negative current region
    discharge_mask = I < -min_pulse_current
    if discharge_mask.sum() < 100:
        print(f"  SKIP {os.path.basename(csv_path)}: no discharge pulse found (|I|>{min_pulse_current}A)", flush=True)
        return None
    
    # Find end of discharge pulse (last index where I < -threshold)
    discharge_indices = np.where(discharge_mask)[0]
    pulse_end_idx = discharge_indices[-1] + 1
    
    if pulse_end_idx >= n - 1000:
        print(f"  SKIP {os.path.basename(csv_path)}: discharge pulse too close to end of file", flush=True)
        return None
    
    # Relaxation starts at pulse_end_idx
    # Skip first 10ms for settling (10 samples at 1kHz)
    relax_start = pulse_end_idx + 10
    relax_end = n  # Use all remaining data
    
    if relax_end - relax_start < 1000:
        print(f"  SKIP {os.path.basename(csv_path)}: relaxation too short ({(relax_end-relax_start)/1000:.1f}s)", flush=True)
        return None
    
    relax_t = t[relax_start:relax_end] - t[relax_start]
    relax_V = U[relax_start:relax_end]
    
    # Check current is near zero during relaxation
    relax_I = I[relax_start:relax_end]
    if np.abs(relax_I).mean() > 0.1:
        print(f"  WARNING {os.path.basename(csv_path)}: mean |I| during relax = {np.abs(relax_I).mean():.3f}A", flush=True)
    
    # Compute overpotential: η(t) = -(V(t) - V_end)
    # Post-discharge: voltage recovers UPWARD, so V(t) < V_end
    # Flip sign so η is positive and decaying to 0 (standard LGN input)
    V_end = relax_V[-1]
    eta = -(relax_V - V_end)  # Positive, decaying
    
    # Parse cell info from filename
    basename = os.path.basename(csv_path)
    mcu = None
    for part in basename.split('_'):
        if part.startswith('MCU'):
            mcu = part
            break
    
    # Discharge pulse characteristics
    discharge_I_mean = I[discharge_mask].mean()
    soc_voltage = U[discharge_indices[0]]  # Voltage at start of discharge
    
    return {
        't': relax_t,
        'V': relax_V,
        'eta': eta,
        'V_end': V_end,
        'cell_id': basename,
        'mcu': mcu,
        'n_points': len(relax_t),
        'relax_duration_s': relax_t[-1],
        'V_recovery_mV': (relax_V[-1] - relax_V[0]) * 1000,
        'discharge_I': discharge_I_mean,
        'soc_voltage': soc_voltage,
        'csv_path': csv_path,
    }


# ============================================================================
# Step 2: Curve fit baseline (sum of exponentials)
# ============================================================================

def exp2_model(t, a1, tau1, a2, tau2):
    """Two-exponential decay: η(t) = a1*exp(-t/τ1) + a2*exp(-t/τ2)"""
    return a1 * np.exp(-t / tau1) + a2 * np.exp(-t / tau2)


def fit_curve(t, eta, window_s=None):
    """
    Fit sum-of-two-exponentials to relaxation data.
    Returns tau1, tau2, fit_nrmse or None on failure.
    """
    if window_s is not None:
        mask = t <= window_s
        t = t[mask]
        eta = eta[mask]
    
    # Downsample for speed (keep 1000 points max)
    if len(t) > 1000:
        step = len(t) // 1000
        t = t[::step]
        eta = eta[::step]
    
    # Initial guesses
    eta0 = eta[0]
    if eta0 < 1e-6:
        return None
    bounds = ([0, 0.01, 0, 0.1], [eta0*10, 100, eta0*10, 5000])
    p0 = [eta0*0.5, 1.0, eta0*0.5, 50.0]
    
    try:
        popt, _ = curve_fit(exp2_model, t, eta, p0=p0, bounds=bounds, maxfev=5000)
        a1, tau1, a2, tau2 = popt
        
        # Ensure tau1 < tau2 (fast, slow)
        if tau1 > tau2:
            a1, tau1, a2, tau2 = a2, tau2, a1, tau1
        
        # Fit quality
        eta_pred = exp2_model(t, *popt)
        nrmse = np.sqrt(np.mean((eta - eta_pred)**2)) / (np.max(eta) - np.min(eta) + 1e-12)
        
        return {'tau1': tau1, 'tau2': tau2, 'a1': a1, 'a2': a2, 'nrmse': nrmse}
    except Exception as e:
        return None


# ============================================================================
# Step 3: LGN fitting (requires existing lgn_battery_exp code)
# ============================================================================

def fit_lgn(t, eta, window_s=None, device='cpu'):
    """
    Fit LGN-SD to relaxation data.
    Uses same approach as Stanford pipeline.
    Returns tau1, tau2, A_matrix, nrmse.
    """
    try:
        import torch
        # Import from existing pipeline
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from lgn_battery_exp_core import fit_lgn_sd  # Your existing LGN fitting function
    except ImportError:
        # Fallback: use the inline version
        return fit_lgn_inline(t, eta, window_s, device)
    
    return fit_lgn_inline(t, eta, window_s, device)


def fit_lgn_inline(t, eta, window_s=None, device='cpu'):
    """
    Inline LGN-SD fitting. Self-contained.
    """
    try:
        import torch
        import torch.nn as nn
    except ImportError:
        print("  PyTorch not available, skipping LGN fit", flush=True)
        return None
    
    if window_s is not None:
        mask = t <= window_s
        t = t[mask]
        eta = eta[mask]
    
    # Downsample to ~500 points for speed
    if len(t) > 500:
        step = len(t) // 500
        t = t[::step]
        eta = eta[::step]
    
    # Normalize
    eta0 = eta[0]
    if abs(eta0) < 1e-8:
        return None
    eta_norm = eta / eta0
    
    t_tensor = torch.tensor(t, dtype=torch.float32, device=device)
    y_tensor = torch.tensor(eta_norm, dtype=torch.float32, device=device)
    
    # LGN-SD: A = S - D, S skew-symmetric, D positive semi-definite
    # For 2D: S = [[0, s], [-s, 0]], D = diag(d1, d2) with d1,d2 > 0
    # State: x(t) = exp(At) * x0, output: y(t) = C^T x(t)
    
    class LGN_SD_2D(nn.Module):
        def __init__(self):
            super().__init__()
            self.s = nn.Parameter(torch.tensor(0.0))
            self.d1_raw = nn.Parameter(torch.tensor(0.0))  # softplus -> d1 > 0
            self.d2_raw = nn.Parameter(torch.tensor(-2.0))  # softplus -> d2 > 0 (slow)
            self.c = nn.Parameter(torch.tensor([1.0, 1.0]))
            self.x0 = nn.Parameter(torch.tensor([1.0, 0.0]))
        
        def forward(self, t):
            d1 = torch.nn.functional.softplus(self.d1_raw)
            d2 = torch.nn.functional.softplus(self.d2_raw)
            s = self.s
            
            A = torch.stack([
                torch.stack([-d1, s]),
                torch.stack([-s, -d2])
            ])
            
            # For each time point, compute exp(At) @ x0
            outputs = []
            for ti in t:
                eAt = torch.matrix_exp(A * ti)
                xt = eAt @ self.x0
                yt = self.c @ xt
                outputs.append(yt)
            
            return torch.stack(outputs)
        
        def get_taus(self):
            with torch.no_grad():
                d1 = torch.nn.functional.softplus(self.d1_raw)
                d2 = torch.nn.functional.softplus(self.d2_raw)
                s = self.s
                A = torch.stack([
                    torch.stack([-d1, s]),
                    torch.stack([-s, -d2])
                ])
                eigenvalues = torch.linalg.eigvals(A)
                taus = -1.0 / eigenvalues.real
                taus_sorted = torch.sort(taus.abs())[0]
                return taus_sorted.cpu().numpy()
    
    model = LGN_SD_2D().to(device)
    optimizer = torch.optim.Adam(model.parameters(), lr=0.01)
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=300)
    
    best_loss = float('inf')
    best_state = None
    
    for epoch in range(300):
        optimizer.zero_grad()
        y_pred = model(t_tensor)
        loss = torch.mean((y_pred - y_tensor)**2)
        loss.backward()
        optimizer.step()
        scheduler.step()
        
        if loss.item() < best_loss:
            best_loss = loss.item()
            best_state = {k: v.clone() for k, v in model.state_dict().items()}
    
    if best_state is not None:
        model.load_state_dict(best_state)
    
    taus = model.get_taus()
    
    # Compute NRMSE
    with torch.no_grad():
        y_pred = model(t_tensor).cpu().numpy()
    nrmse = np.sqrt(np.mean((eta_norm - y_pred)**2)) / (np.max(eta_norm) - np.min(eta_norm) + 1e-12)
    
    return {
        'tau1': float(min(taus)),
        'tau2': float(max(taus)),
        'nrmse': float(nrmse),
    }


# ============================================================================
# Step 4: Parse Batch_List for SOH labels
# ============================================================================

def load_batch_list(xlsx_path):
    """
    Parse Batch_List.xlsx to get SOH category for each cell.
    Returns dict: batch_number -> {position -> soh_info}
    """
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not available, install with: pip install openpyxl", flush=True)
        return None
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Daten']
    
    cells_info = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        number = row[0]
        soh_raw = row[8] if row[8] else 'Unknown'
        batch = row[9]
        position = row[10]
        voltage = row[4]
        quality = row[6]
        
        # Parse SOH category
        if '100' in str(soh_raw):
            soh_category = 'SoH>95%'
        elif '90' in str(soh_raw) and '95' in str(soh_raw):
            soh_category = 'SoH 90-95%'
        elif 'neu' in str(soh_raw).lower():
            soh_category = 'New'
        else:
            soh_category = 'Unknown'
        
        # Parse Z@1kHz if available
        z_1khz = None
        soh_str = str(soh_raw)
        if 'mOhm' in soh_str or 'mOhm' in soh_str:
            import re
            z_match = re.search(r'(\d+[\.,]\d+)\s*mOhm', soh_str)
            if z_match:
                z_1khz = float(z_match.group(1).replace(',', '.'))
        
        cells_info.append({
            'number': number,
            'batch': batch,
            'position': position,
            'soh_raw': soh_raw,
            'soh_category': soh_category,
            'z_1khz_supplier': z_1khz,
            'voltage': voltage,
            'quality': quality,
        })
    
    return cells_info


# ============================================================================
# Step 5: Main pipeline
# ============================================================================

def discover_csv_files(data_dir):
    """Find all Popp CSV files, organized by batch."""
    patterns = [
        os.path.join(data_dir, '**', 'Engineered_Cleaned_Report_*.csv'),
        os.path.join(data_dir, 'Engineered_Cleaned_Report_*.csv'),
        os.path.join(data_dir, '**', '*.csv'),
    ]
    all_files = []
    for pat in patterns:
        files = glob.glob(pat, recursive=True)
        all_files.extend(files)
    
    # Deduplicate
    all_files = sorted(set(all_files))
    
    # Filter for HPPC files (contain MCU in name)
    hppc_files = [f for f in all_files if 'MCU' in os.path.basename(f)]
    
    if not hppc_files:
        # Try all CSVs
        hppc_files = all_files
    
    print(f"Found {len(hppc_files)} HPPC CSV files", flush=True)
    return hppc_files


def run_pipeline(args):
    print(f"{'='*70}", flush=True)
    print(f"POPP 256-CELL LGN PIPELINE", flush=True)
    print(f"{'='*70}", flush=True)
    
    # Discover files
    if args.file_list and os.path.exists(args.file_list):
        with open(args.file_list) as f:
            csv_files = [line.strip() for line in f if line.strip()]
        print(f"Loaded {len(csv_files)} files from {args.file_list}", flush=True)
    else:
        csv_files = discover_csv_files(args.data_dir)
    if not csv_files:
        print("ERROR: No CSV files found!", flush=True)
        return
    
    if args.max_cells:
        csv_files = csv_files[:args.max_cells]
    
    # Load batch list for SOH labels
    cells_info = None
    if args.batch_list and os.path.exists(args.batch_list):
        cells_info = load_batch_list(args.batch_list)
        print(f"Loaded {len(cells_info)} cell records from Batch_List", flush=True)
    
    # Process each cell
    results = []
    device = f'cuda:{args.gpu}' if args.gpu >= 0 else 'cpu'
    
    print(f"\nProcessing {len(csv_files)} cells on {device}...", flush=True)
    print(f"Windows: full, w100 (first 10s), w300 (first 30s)", flush=True)
    
    for fi, csv_path in enumerate(csv_files):
        basename = os.path.basename(csv_path)
        print(f"\n[{fi+1}/{len(csv_files)}] {basename}", flush=True)
        
        # Step 1: Extract relaxation
        relax = extract_relaxation(csv_path)
        if relax is None:
            continue
        
        print(f"  Relaxation: {relax['relax_duration_s']:.1f}s, "
              f"{relax['n_points']} pts, "
              f"ΔV={relax['V_recovery_mV']:.1f} mV")
        
        t = relax['t']
        eta = relax['eta']
        
        # Step 2: Fit with multiple windows
        result = {
            'cell_id': relax['cell_id'],
            'mcu': relax['mcu'],
            'csv_path': csv_path,
            'V_end': relax['V_end'],
            'V_recovery_mV': relax['V_recovery_mV'],
            'discharge_I': relax['discharge_I'],
            'soc_voltage': relax['soc_voltage'],
            'relax_duration_s': relax['relax_duration_s'],
        }
        
        windows = {'full': None, 'w100': 10.0, 'w300': 30.0}
        
        for wname, wlen in windows.items():
            # Curve fit baseline
            cf = fit_curve(t, eta, window_s=wlen)
            if cf:
                result[f'cf_tau1_{wname}'] = cf['tau1']
                result[f'cf_tau2_{wname}'] = cf['tau2']
                result[f'cf_nrmse_{wname}'] = cf['nrmse']
                print(f"  CF  {wname}: τ₁={cf['tau1']:.3f}s, τ₂={cf['tau2']:.1f}s, NRMSE={cf['nrmse']:.4f}", flush=True)
            
            # LGN fit
            if not args.cf_only:
                lgn = fit_lgn_inline(t, eta, window_s=wlen, device=device)
                if lgn:
                    result[f'lgn_tau1_{wname}'] = lgn['tau1']
                    result[f'lgn_tau2_{wname}'] = lgn['tau2']
                    result[f'lgn_nrmse_{wname}'] = lgn['nrmse']
                    print(f"  LGN {wname}: τ₁={lgn['tau1']:.3f}s, τ₂={lgn['tau2']:.1f}s, NRMSE={lgn['nrmse']:.4f}", flush=True)
        
        # Compute R_pulse = total voltage recovery / discharge current magnitude
        if abs(relax['discharge_I']) > 0.1:
            result['R_pulse'] = abs(relax['V_recovery_mV'] / 1000 / relax['discharge_I'])
        
        results.append(result)
    
    print(f"\n{'='*70}", flush=True)
    print(f"Processed {len(results)} / {len(csv_files)} cells successfully", flush=True)
    
    # ---- Match with SOH labels ----
    # For now, assign SOH based on folder/batch structure
    # The Batch_List maps cell numbers to SOH categories
    # MCU files map to positions within a batch
    
    # Save raw results
    os.makedirs(args.out_dir, exist_ok=True)
    
    # Convert numpy types for JSON serialization
    clean_results = []
    for r in results:
        clean_r = {}
        for k, v in r.items():
            if isinstance(v, (np.floating, np.integer)):
                clean_r[k] = float(v)
            elif isinstance(v, np.ndarray):
                clean_r[k] = v.tolist()
            else:
                clean_r[k] = v
        clean_results.append(clean_r)
    
    with open(f'{args.out_dir}/popp_results.json', 'w') as f:
        json.dump(clean_results, f, indent=2)
    print(f"\nSaved to {args.out_dir}/popp_results.json", flush=True)
    
    # ---- Statistical analysis ----
    if len(results) > 10:
        analyze_results(results, cells_info, args.out_dir, args.plot)


def analyze_results(results, cells_info, out_dir, do_plot):
    """Statistical analysis: does τ separate SOH groups?"""
    print(f"\n{'='*70}", flush=True)
    print(f"STATISTICAL ANALYSIS", flush=True)
    print(f"{'='*70}", flush=True)
    
    df = pd.DataFrame(results)
    
    # Key features to analyze
    tau_cols = [c for c in df.columns if 'tau' in c and df[c].notna().sum() > 5]
    
    print(f"\nAvailable τ features: {tau_cols}", flush=True)
    print(f"Cells with R_pulse: {df['R_pulse'].notna().sum() if 'R_pulse' in df.columns else 0}", flush=True)
    
    # Correlation between τ and V_end (proxy for SOH - healthier cells have higher OCV)
    print(f"\n--- Correlation: τ vs V_end (OCV proxy for SOH) ---", flush=True)
    for col in tau_cols:
        valid = df[[col, 'V_end']].dropna()
        if len(valid) > 5:
            rho, p = stats.spearmanr(valid[col], valid['V_end'])
            sig = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else 'n.s.'
            print(f"  {col:25s}: ρ = {rho:+.3f}  (p = {p:.2e}) {sig}", flush=True)
    
    if 'R_pulse' in df.columns:
        valid = df[['R_pulse', 'V_end']].dropna()
        if len(valid) > 5:
            rho, p = stats.spearmanr(valid['R_pulse'], valid['V_end'])
            sig = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else 'n.s.'
            print(f"  {'R_pulse':25s}: ρ = {rho:+.3f}  (p = {p:.2e}) {sig}", flush=True)
    
    # Correlation: τ vs V_recovery (larger recovery = more degraded)
    print(f"\n--- Correlation: τ vs V_recovery_mV (degradation proxy) ---", flush=True)
    for col in tau_cols:
        valid = df[[col, 'V_recovery_mV']].dropna()
        if len(valid) > 5:
            rho, p = stats.spearmanr(valid[col], valid['V_recovery_mV'])
            sig = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else 'n.s.'
            print(f"  {col:25s}: ρ = {rho:+.3f}  (p = {p:.2e}) {sig}", flush=True)
    
    # Summary statistics
    print(f"\n--- τ summary statistics ---", flush=True)
    for col in tau_cols:
        vals = df[col].dropna()
        print(f"  {col:25s}: mean={vals.mean():.3f}  std={vals.std():.3f}  "
              f"range=[{vals.min():.3f}, {vals.max():.3f}]  n={len(vals)}")
    
    if do_plot:
        plot_popp_results(df, tau_cols, out_dir)


def plot_popp_results(df, tau_cols, out_dir):
    """Publication-quality plots for Popp results."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    
    os.makedirs(out_dir, exist_ok=True)
    plt.rcParams.update({'font.size': 11, 'savefig.dpi': 300, 'savefig.bbox': 'tight'})
    
    # --- Figure 1: τ distributions ---
    fig, axes = plt.subplots(1, min(4, len(tau_cols)), figsize=(4*min(4, len(tau_cols)), 4))
    if len(tau_cols) == 1:
        axes = [axes]
    for i, col in enumerate(tau_cols[:4]):
        ax = axes[i]
        vals = df[col].dropna()
        ax.hist(vals, bins=30, color='steelblue', edgecolor='black', alpha=0.7)
        ax.set_xlabel(col)
        ax.set_ylabel('Count')
        ax.set_title(f'n={len(vals)}')
    fig.suptitle('τ Distributions Across Popp Cells', fontweight='bold')
    fig.tight_layout()
    fig.savefig(f'{out_dir}/popp_tau_distributions.png')
    plt.close(fig)
    
    # --- Figure 2: τ vs V_end scatter ---
    n_plots = min(4, len(tau_cols))
    fig, axes = plt.subplots(1, n_plots, figsize=(4*n_plots, 4))
    if n_plots == 1:
        axes = [axes]
    for i, col in enumerate(tau_cols[:4]):
        ax = axes[i]
        valid = df[[col, 'V_end']].dropna()
        ax.scatter(valid['V_end'], valid[col], s=15, alpha=0.5)
        rho, p = stats.spearmanr(valid['V_end'], valid[col])
        ax.set_xlabel('V_end (OCV) [V]')
        ax.set_ylabel(col)
        ax.set_title(f'ρ={rho:.3f}, p={p:.1e}')
        ax.grid(True, alpha=0.2)
    fig.suptitle('τ vs Open Circuit Voltage (SOH Proxy)', fontweight='bold')
    fig.tight_layout()
    fig.savefig(f'{out_dir}/popp_tau_vs_soh.png')
    plt.close(fig)
    
    # --- Figure 3: τ vs V_recovery scatter ---
    fig, axes = plt.subplots(1, n_plots, figsize=(4*n_plots, 4))
    if n_plots == 1:
        axes = [axes]
    for i, col in enumerate(tau_cols[:4]):
        ax = axes[i]
        valid = df[[col, 'V_recovery_mV']].dropna()
        ax.scatter(valid['V_recovery_mV'], valid[col], s=15, alpha=0.5, color='coral')
        rho, p = stats.spearmanr(valid['V_recovery_mV'], valid[col])
        ax.set_xlabel('V recovery [mV]')
        ax.set_ylabel(col)
        ax.set_title(f'ρ={rho:.3f}, p={p:.1e}')
        ax.grid(True, alpha=0.2)
    fig.suptitle('τ vs Voltage Recovery (Degradation Proxy)', fontweight='bold')
    fig.tight_layout()
    fig.savefig(f'{out_dir}/popp_tau_vs_recovery.png')
    plt.close(fig)
    
    # --- Figure 4: LGN vs CF comparison (if both available) ---
    lgn_cols = [c for c in tau_cols if 'lgn' in c]
    cf_cols = [c for c in tau_cols if 'cf' in c]
    if lgn_cols and cf_cols:
        fig, axes = plt.subplots(1, 2, figsize=(10, 5))
        for i, suffix in enumerate(['tau1', 'tau2']):
            ax = axes[i]
            lgn_col = f'lgn_{suffix}_w300'
            cf_col = f'cf_{suffix}_w300'
            if lgn_col in df.columns and cf_col in df.columns:
                valid = df[[lgn_col, cf_col]].dropna()
                ax.scatter(valid[cf_col], valid[lgn_col], s=15, alpha=0.5)
                lims = [min(valid[cf_col].min(), valid[lgn_col].min()),
                        max(valid[cf_col].max(), valid[lgn_col].max())]
                ax.plot(lims, lims, 'k--', alpha=0.3)
                ax.set_xlabel(f'Curve Fit {suffix}')
                ax.set_ylabel(f'LGN {suffix}')
                rho, _ = stats.spearmanr(valid[cf_col], valid[lgn_col])
                ax.set_title(f'{suffix}: ρ={rho:.3f}')
                ax.grid(True, alpha=0.2)
        fig.suptitle('LGN vs Curve Fit: τ Comparison', fontweight='bold')
        fig.tight_layout()
        fig.savefig(f'{out_dir}/popp_lgn_vs_cf.png')
        plt.close(fig)
    
    print(f"  Plots saved to {out_dir}/", flush=True)


# ============================================================================
if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('--data_dir', required=True,
                   help='Path to folder containing Popp CSV files')
    p.add_argument('--batch_list', default='Batch_List.xlsx',
                   help='Path to Batch_List.xlsx')
    p.add_argument('--out_dir', default='results_popp')
    p.add_argument('--gpu', type=int, default=0,
                   help='GPU index (-1 for CPU)')
    p.add_argument('--cf_only', action='store_true',
                   help='Only run curve fit (skip LGN, faster)')
    p.add_argument('--plot', action='store_true')
    p.add_argument('--max_cells', type=int, default=None,
                   help='Max cells to process (for testing)')
    p.add_argument('--file_list', default=None,
                   help='Text file with one CSV path per line (for parallel runs)')
    args = p.parse_args()
    
    run_pipeline(args)
